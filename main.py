import asyncio
import datetime
import re
import csv
import os
from sys import path
from openpyxl import load_workbook
from selenium import webdriver


from selenium.common import NoSuchElementException, WebDriverException
from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


RAW_DATA_PATH = "USA Services.xlsx"
PATH_TO_PROCESSED_CSV = "csv"


chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument(
    "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    " (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"
)


async def open_web_driver():
    driver = webdriver.Chrome(options=chrome_options)
    return driver


async def close_web_driver(driver):
    driver.quit()


async def writing_to_csv(sheet_name, data):
    csv_folder = "csv"
    if not os.path.exists(csv_folder):
        os.makedirs(csv_folder)

    csv_file_path = os.path.join(csv_folder, f'{sheet_name}.csv')

    with open(csv_file_path, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(csvfile)
        headers = [
            "Keyword", "Location", "Name", "Website", "Phone",
            "Instagram", "Facebook", "Twitter", "Additional Phone", "Email",
        ]
        writer.writerow(headers)

        for row_data in data:
            writer.writerow(row_data)


async def find_email_on_facebook(driver, url):
    driver.get(url)
    try:
        button = await WebDriverWait(driver, 4).until(
            EC.element_to_be_clickable(
                (
                    By.CSS_SELECTOR,
                    '[role="button"][tabindex="0"] > [data-visualcompletion="css-img"]',
                )
            )
        )
        button.click()
    except Exception:
        pass

    WebDriverWait(driver, 2).until(
        EC.presence_of_element_located((By.TAG_NAME, "body"))
    )
    emails = re.findall(
        r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b",
        driver.page_source,
    )
    return emails


async def simple_captcha_click(driver):
    try:
        challenge_stage = driver.find_element(By.ID, "challenge-stage")
        checkbox = challenge_stage.find_element(By.CSS_SELECTOR, 'input[type="checkbox"]')
        checkbox.click()
        print("Captcha was clicked")

    except NoSuchElementException:
        pass


async def get_social_links_and_additional_number(driver):
    instagrams = set()
    facebooks = set()
    twitters = set()
    phone_number = set()

    social_links = driver.find_elements(By.TAG_NAME, "a")
    if len(social_links) >= 1:
        for link in social_links:
            href = link.get_attribute('href')
            if href is not None:
                if "facebook" in href:
                    facebooks.add(href)
                if "twitter" in href:
                    twitters.add(href)
                if "instagram" in href:
                    instagrams.add(href)
                if "tel:" in href:
                    phone_number.add(href.split(":")[1])

    return instagrams, facebooks, twitters, phone_number


async def get_email_address(driver, page_text, facebooks):
    email_address = None
    try:
        emails = re.findall(
            r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b", page_text
        )
        if len(emails) >= 1:
            email_address = set(emails)
        else:
            for fb_url in facebooks:
                email = await find_email_on_facebook(driver, fb_url)
                if email:
                    email_address = email[0]
                    break
    except Exception as e:
        print(f"Error getting email address: {e}")
    return email_address


async def parse_single_website(driver, website: str):
    print(f"Start parsing site: {website}")

    try:
        driver.get("https://" + website)

        await simple_captcha_click(driver)

        WebDriverWait(driver, 2).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        instagrams, facebooks, twitters, phone_number = await get_social_links_and_additional_number(driver)

        page_text = driver.page_source
        email_address = await get_email_address(driver, page_text, facebooks)

    except WebDriverException:
        print(f"Error parsing website {website}")
        return None, None, None, None, None

    print(f"Finish parsing site: {website}")
    return instagrams, facebooks, twitters, phone_number, email_address


async def async_organizing_excel_data(file_path: path, driver: WebDriver):
    workbook = load_workbook(file_path)
    sheets = workbook.sheetnames

    for sheet_name in sheets[:1]:
        sheet = workbook[sheet_name]
        data = []

        for row in sheet.iter_rows(
            min_row=40, max_row=45, min_col=1, max_col=5, values_only=True
        ):
            keyword, location, name, websites, phone = row

            if re.match(r"\+\d{1,3} \d{3}-\d{3}-\d{4}", str(websites)):
                phone, websites = websites, phone

            try:
                websites = [site.strip() for site in websites.split(",")]

                for website in set(websites):
                    instagrams, facebooks, twitters, phone_numbers, email_address = (
                        await parse_single_website(driver, website)
                    )

                    instagrams_str = ", ".join(instagrams) if instagrams else ""
                    facebooks_str = ", ".join(facebooks) if facebooks else ""
                    twitters_str = ", ".join(twitters) if twitters else ""
                    phone_numbers_str = ", ".join(phone_numbers) if phone_numbers else ""

                    new_row = (
                        keyword, location, name, website, phone, instagrams_str,
                        facebooks_str, twitters_str, phone_numbers_str, email_address,
                    )
                    data.append(new_row)
            except AttributeError:
                pass
        await writing_to_csv(sheet_name, data)


async def main():
    start = datetime.datetime.now()
    driver = await open_web_driver()
    try:
        await async_organizing_excel_data(RAW_DATA_PATH, driver)
    finally:
        await close_web_driver(driver)
    print("Total time is: ", datetime.datetime.now() - start)

if __name__ == "__main__":
    asyncio.run(main())
