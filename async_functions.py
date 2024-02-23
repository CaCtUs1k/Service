import datetime
import json
import re
import csv
import os
import aiofiles
import httpx

from httpx import Response, AsyncClient
from ssl import SSLError
from anyio import EndOfStream
from openpyxl import load_workbook

# Load configuration from a JSON file
with open("config.json", "r") as file:
    config = json.load(file)
    PROCESSED_FILES_DIR = config["PROCESSED_FILES_DIR"]
    RAW_DATA_PATH = config["RAW_DATA_PATH"]
    SEMIFINISHED_FILES_DIR = config["SEMIFINISHED_FILES_DIR"]


async def writing_to_csv(sheet_name: str, data: list) -> None:
    """
    Write data to a CSV file.
    """

    # Create the directory for storing CSV files if it doesn't exist
    if not os.path.exists(SEMIFINISHED_FILES_DIR):
        os.makedirs(SEMIFINISHED_FILES_DIR)

    # Construct the file path for the CSV file
    csv_file_path = os.path.join(SEMIFINISHED_FILES_DIR, f"{sheet_name}.csv")

    # Open the CSV file in asynchronous mode
    async with aiofiles.open(
        csv_file_path, mode="w", encoding="utf-8", newline=""
    ) as csvfile:
        writer = csv.writer(csvfile)

        # Write headers to the CSV file
        headers = [
            "Keyword",
            "Location",
            "Name",
            "Website",
            "Phone",
            "Instagram",
            "Facebook",
            "Twitter",
            "Additional Phone",
            "Email",
        ]
        await writer.writerow(headers)

        # Write each row of data to the CSV file
        for row_data in data:
            await writer.writerow(row_data)


async def get_email_address(response: Response) -> str:
    """
    Extract email address from HTTP response text.

    Returns:
        str: Extracted email address.
    """
    email_address = None
    try:
        # Use regular expression to find email address in response text
        match = re.search(
            r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b",
            response.text,
        )
        if match:
            email_address = match.group()
    except Exception as e:
        print(f"Error getting email address: {e}")
    return email_address


async def get_social_links_and_additional_number(response: Response) -> tuple:
    """
    Extract social links and additional phone numbers from HTTP response text.

    Returns:
        tuple: Tuple containing strings of Instagram links, Facebook links, Twitter links,
               and additional phone numbers.
    """
    instagrams = set()
    facebooks = set()
    twitters = set()
    phone_number = set()

    try:
        # Use regular expression to find social links and phone numbers in response text
        social_links = re.findall(
            r'<a\s+(?:[^>]*?\s+)?href="([^"]*)"', response.text
        )

        for link in social_links:
            if "facebook" in link:
                facebooks.add(link)
            elif "twitter" in link:
                twitters.add(link)
            elif "instagram" in link:
                instagrams.add(link)
            elif link.startswith("tel:"):
                phone_number.add(link.split(":")[1])

    except (httpx.HTTPStatusError, httpx.RequestError) as exc:
        print(f"Error getting social links and additional number: {exc}")

    # Convert sets to strings
    instagrams_str = ", ".join(instagrams) if instagrams else ""
    facebooks_str = ", ".join(facebooks) if facebooks else ""
    twitters_str = ", ".join(twitters) if twitters else ""
    phone_numbers_str = ", ".join(phone_number) if phone_number else ""

    return instagrams_str, facebooks_str, twitters_str, phone_numbers_str


async def parse_single_website(client: AsyncClient, website: str) -> tuple:
    """
    Parse data from a single website.

    Returns:
        tuple: Tuple containing strings of Instagram links, Facebook links, Twitter links,
               additional phone numbers, and email address.
    """
    print(f"Start parsing website: {website}")
    try:
        # Send an HTTP GET request to the website
        response = await client.get(website)

        # Handle redirection if necessary
        if response.status_code in [301, 302, 308] and response.headers.get(
            "location"
        ) not in [website, None]:
            new_website = response.headers.get("location")
            return await parse_single_website(client, new_website)

        # If response is successful (status code 200), extract data
        if response.status_code == 200:
            instagrams, facebooks, twitters, phone_number = (
                await get_social_links_and_additional_number(response)
            )
            emails = await get_email_address(response)
            return instagrams, facebooks, twitters, phone_number, emails

        response.raise_for_status()

    except (httpx.HTTPStatusError, httpx.RequestError, ValueError) as exc:
        print(f"Error parsing website {website}: {exc}")
        return None, None, None, None, None


async def organizing_excel_data(file_path: str, client: AsyncClient) -> None:
    """
    Organize data from an Excel file.
    """

    # Load the Excel workbook
    workbook = load_workbook(file_path, read_only=True)
    sheets = workbook.sheetnames

    # Iterate through each sheet in the workbook
    for sheet_name in sheets:
        start = datetime.datetime.now()
        sheet = workbook[sheet_name]
        data = []
        try:
            for row in sheet.iter_rows(
                min_row=2, min_col=1, max_col=5, values_only=True
            ):
                keyword, location, name, websites, phone = row

                # Swap phone and website if phone matches a specific format
                if re.match(r"\+\d{1,3} \d{3}-\d{3}-\d{4}", str(websites)):
                    phone, websites = websites, phone

                try:

                    # If websites exist, parse data from each website
                    if websites:
                        websites = [
                            site.strip() for site in websites.split(",")
                        ]

                        for website in set(websites):

                            # Skip certain websites
                            if (
                                "facebook" in website
                                or website == "business.site"
                            ):
                                continue

                            # Parse data from the website
                            (
                                instagrams,
                                facebooks,
                                twitters,
                                phone_numbers,
                                email_address,
                            ) = await parse_single_website(
                                client, "https://" + website
                            )

                            new_row = (
                                keyword,
                                location,
                                name,
                                website,
                                phone,
                                instagrams,
                                facebooks,
                                twitters,
                                phone_numbers,
                                email_address,
                            )
                            data.append(new_row)
                except (AttributeError, EndOfStream, SSLError) as e:
                    print(e)
                    pass
        finally:
            print(f"Sheet {sheet_name} was parsed")
            print(
                f"{sheet_name} parsing time is: ",
                datetime.datetime.now() - start,
            )
            # Write parsed data to a CSV file
            await writing_to_csv(sheet_name, data)
